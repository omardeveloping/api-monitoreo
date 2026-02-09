import io

from rest_framework.exceptions import ValidationError

from dashboard.services.importar_velocidades_csv import importar_velocidades_tabulares


def importar_velocidades_xlsx(video, archivo):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("openpyxl no está instalado en el servidor.") from exc

    contenido = archivo.read()
    if not contenido:
        raise ValidationError("El XLSX está vacío.")

    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    filas_iter = ws.iter_rows(values_only=True)

    try:
        encabezados_raw = next(filas_iter)
    except StopIteration as exc:
        raise ValidationError("El XLSX no tiene encabezados.") from exc

    encabezados = [
        (str(valor).strip() if valor is not None else "")
        for valor in encabezados_raw
    ]

    if not any(encabezados):
        raise ValidationError("El XLSX no tiene encabezados válidos.")

    filas = (
        {
            encabezados[idx]: valor
            for idx, valor in enumerate(fila)
            if idx < len(encabezados)
        }
        for fila in filas_iter
    )

    return importar_velocidades_tabulares(video, encabezados, filas)
